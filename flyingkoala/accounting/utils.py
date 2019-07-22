# utils.py
import xlwings as xw
from xlwings import constants
from openpyxl import utils


def delete_columns(sheet_to_delete_from, columns_to_keep, sheet_rename, delete=True):
    if not isinstance(columns_to_keep, list):
        return

    # get the worksheet we are interested in (providing it exists)
    wb = xw.Book.caller()
    sheets = xw.sheets
    working_sheet = None
    for sheet in sheets:
        if sheet.name == sheet_to_delete_from:
            working_sheet = sheet

    if working_sheet is not None:
        columns = working_sheet.used_range.columns.count

        # excel column references are 1 based
        columns += 1
        length = 0
        column_ordinals_to_delete = []

        for col in range(1, columns):
            length += 1
            # are we getting rid of these columns, or keeping them and deleting the others..?
            if delete:
                if working_sheet.cells(1, col).value not in columns_to_keep:
                    column_ordinals_to_delete.append(length)
            else:
                if working_sheet.cells(1, col).value in columns_to_keep:
                    column_ordinals_to_delete.append(length)

        # delete from the right-hand-side first so the ordinals don't change.
        column_ordinals_to_delete.reverse()

        for column in column_ordinals_to_delete:
            letter = utils.cell.get_column_letter(column)
            working_sheet.range('%s:%s' % (letter, letter)).api.Delete(constants.DeleteShiftDirection.xlShiftToLeft)

        working_sheet.name = sheet_rename

def keep_columns(sheet_to_delete_from, columns_to_delete, sheet_rename):
    return delete_columns(sheet_to_delete_from, columns_to_delete, sheet_rename, delete=False)


def employee_rate_and_total(sheet_range = 'Employee Hours', sheet_lookup = 'Employee Rates'):

    # get the worksheet we are interested in (providing it exists)
    wb = xw.Book.caller()
    sheets = xw.sheets
    working_sheet = None
    for sheet in sheets:
        if sheet.name == sheet_range:
            working_sheet = sheet

    if working_sheet is not None:
        Lookup_cols = working_sheet.used_range.columns.count
        Lookup_rows = working_sheet.used_range.rows.count + 1

        rates_col = Lookup_cols + 1                   #calculates the column number for the rates
        total_col = Lookup_cols + 2                    #calculates the row number for the rates
        rows = working_sheet.used_range.rows.count             #calculates the number of rows in "Sheet1"
        code_position = 0
        surname_position = 0
        hours_position = 0

        for col in range(1, Lookup_cols + 1):
            if working_sheet.cells(1, col).value == 'Project Code':
                code_position = col

            if working_sheet.cells(1, col).value == 'Last Name':
                surname_position = col

            if working_sheet.cells(1, col).value == 'Hours':
                hours_position = col

        rates_Letter = utils.cell.get_column_letter(rates_col)
        total_Letter = utils.cell.get_column_letter(total_col)
        code_Letter = utils.cell.get_column_letter(code_position)
        surname_Letter = utils.cell.get_column_letter(surname_position)
        hours_Letter = utils.cell.get_column_letter(hours_position)

        working_sheet.cells(1,rates_col).value = "rates"
        working_sheet.cells(2,rates_col).value = "=VLOOKUP(%s%s, '%s'!$A$1:$B$%s, 2, FALSE)" % (surname_Letter, 2, sheet_lookup, Lookup_rows)
        working_sheet.cells(2,rates_col).api.autofill(working_sheet.range("%s2:%s%s" % (rates_Letter, rates_Letter, rows)).api, constants.AutoFillType.xlFillDefault)

        working_sheet.cells(1,total_col).value = "total"
        working_sheet.cells(2,total_col).value = "=%s2*%s2" %(hours_Letter, rates_Letter)
        working_sheet.cells(2,total_col).api.autofill(working_sheet.range("%s2:%s%s" % (total_Letter, total_Letter, rows)).api, constants.AutoFillType.xlFillDefault)

def make_employee_pivot_table():
    # get the worksheet we are interested in (providing it exists)
    wb = xw.Book.caller()
    sheets = xw.sheets
    working_sheet = None
    for sheet in sheets:
        if sheet.name == 'Employee Hours':
            working_sheet = sheet

    for col in range(1, working_sheet.used_range.columns.count + 1):
        if working_sheet.cells(1, col).value == 'Project Code':
            code_position = col

        if working_sheet.cells(1, col).value == 'Last Name':
            surname_position = col

        if working_sheet.cells(1, col).value == 'total':
            total_position = col

    code_Letter = utils.cell.get_column_letter(code_position)
    surname_Letter = utils.cell.get_column_letter(surname_position)
    total_Letter = utils.cell.get_column_letter(total_position)

    pivot_data = xw.sheets.add(name='Pivot Data', after='Sheet1')

    if working_sheet is not None:
        working_sheet.range("%s:%s" % (code_Letter, code_Letter)).api.Copy(pivot_data.range("A:A").api )
        working_sheet.range("%s:%s" % (surname_Letter, surname_Letter)).api.Copy(pivot_data.range("B:B").api )

        working_sheet.range("%s:%s" % (total_Letter, total_Letter)).api.Copy()
        pivot_data.range("C1").api.PasteSpecial(Paste = constants.PasteType.xlPasteValues)

        val_rows = pivot_data.used_range.rows.count                             #counts the number of rows in the values sheet

        PivotSourceRange = pivot_data.range("A1:C%s" % val_rows)                #selects the source data for the pivot table

        pivot_table = xw.sheets.add(name='Pivot Table', after='Sheet1')

        PivotTableName = 'ReportPivotTable'

        PivotCache = wb.api.PivotCaches().Create(SourceType=constants.PivotTableSourceType.xlDatabase, SourceData=PivotSourceRange.api, Version=constants.PivotTableVersionList.xlPivotTableVersion14)
        PivotTable = PivotCache.CreatePivotTable(TableDestination="'Pivot Table'!R1C1", TableName=PivotTableName, DefaultVersion=constants.PivotTableVersionList.xlPivotTableVersion14)
        PivotTable.PivotFields('Last Name').Orientation = constants.PivotFieldOrientation.xlRowField
        PivotTable.PivotFields('Last Name').Position = 1
        PivotTable.PivotFields('Project Code').Orientation = constants.PivotFieldOrientation.xlRowField
        PivotTable.PivotFields('Project Code').Position = 2
        PivotTable.PivotFields("total").Orientation = constants.PivotFieldOrientation.xlDataField

        piv_rows = pivot_table.used_range.rows.count                             #counts the number of rows in the pivot table

        people = []
        for each in PivotTable.PivotFields("Last Name").PivotItems():     #creates a list containing the names of all the people in the pivot table
            people.append(each.value)

        the_name = None
        for row in range(1, piv_rows):                                   #matches the name to the corresping code
            if (pivot_table.cells(row, 1).value in people):
                the_name = pivot_table.cells(row, 1).value
            else:
                pivot_table.cells(row, 3).value = the_name

        people = ["Row Labels", "Grand Total"] + people                                #adds 'Row Labels' and 'Grand Total' to the start and finish of the people list

        pivot_text = xw.sheets.add(name='Pivot Text', after='Sheet1')
        pivot_text.cells(1, 1).value = 'Project Code'
        pivot_text.cells(1, 2).value = 'total'
        pivot_text.cells(1, 3).value = 'Last Name'

        #copies and pastes the the pivot table to the pivot_text sheet
        pivot_table.range("A1:C%s" % piv_rows).api.Copy()
        pivot_text.range("A2").api.PasteSpecial(Paste = constants.PasteType.xlPasteValues)

        for row in reversed( range(1, piv_rows + 1) ):
            if pivot_text.cells(row, 1).value in people:
                pivot_text.cells(row, 1).api.EntireRow.Delete()

def project_code_and_names(sheet_range = 'Pivot Text', sheet_lookup = 'Project Code and Names'):

    # get the worksheet we are interested in (providing it exists)
    wb = xw.Book.caller()
    sheets = xw.sheets
    working_sheet = None
    for sheet in sheets:
        if sheet.name == sheet_range:
            working_sheet = sheet

    if working_sheet is not None:
        working_sheet.cells(1, 4).value = 'Project Code and Name'

        rows = working_sheet.used_range.rows.count             #calculates the number of rows in "Sheet1"

        working_sheet.cells(2,4).value = "=VLOOKUP(A2, '%s'!$A:$B, 2, FALSE)" % (sheet_lookup)
        working_sheet.cells(2,4).api.autofill(working_sheet.range("$D$2:$D$%s" % rows).api, constants.AutoFillType.xlFillDefault)
